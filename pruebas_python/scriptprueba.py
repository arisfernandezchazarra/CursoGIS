import pandas as pd
from openpyxl import load_workbook
import os

def update_excel_with_extracted_data(excel_path, new_data):
    """
    Inserta una nueva fila de datos en el archivo Excel existente.
    """
    try:
        # Si el archivo no existe, creamos uno nuevo con los datos como primera fila
        if not os.path.exists(excel_path):
            df = pd.DataFrame([new_data])
            # Definimos nombres de columnas genéricos o basados en tu esquema
            df.columns = [
                'NIE', 'NOMBRE', 'APELLIDOS', 'TELEFONO', 'EMAIL',
                'CIF EMPRESA', 'NOMBRE EMPRESA', 'DIRECCION EMPRESA', 'LOCALIDAD EMPRESA', 'TLF EMPRESA',
                'REPRESENTANTE EMPRESA', 'DNI REPRESENTANTE',
                'TUTOR EMPRESA', 'DNI TUTOR', 'FECHA NACIMIENTO', 'CORREO EMPRESA'
            ]
            df.to_excel(excel_path, index=False)
            return f"Archivo nuevo creado: {excel_path}"

        # Si existe, cargamos el libro para añadir la fila al final
        book = load_workbook(excel_path)
        sheet = book.active

        # Buscamos la primera fila vacía
        next_row = sheet.max_row + 1

        # Insertamos los datos de la columna A (1) a la P (16)
        for col_idx, value in enumerate(new_data, start=1):
            sheet.cell(row=next_row, column=col_idx).value = value

        book.save(excel_path)
        return f"Éxito: Datos de {new_data[1]} {new_data[2]} insertados en la fila {next_row}."
    
    except Exception as e:
        return f"Error al procesar el archivo Excel: {str(e)}"

# Datos mapeados según el PDF extraído (Yamina Baroudi)
data_to_insert = [
    "Y1234567A",                      # A - NIE
    "YAMINA",                         # B - NOMBRE
    "BAROUDI",                        # C - APELLIDOS
    "632345821",                      # D - TELEFONO
    "yaminabaroudi@gmail.com",        # E - EMAIL
    "B44545523",                      # F - CIF EMPRESA
    "K KUBIK EXPERIENCE SLU",         # G - NOMBRE EMPRESA
    "C/ SANT ANTONI 23",              # H - DIRECCION EMPRESA
    "SAN JUAN DE ALICANTE (ALICANTE)",# I - LOCALIDAD EMPRESA
    "666824157",                      # J - TLF EMPRESA
    "ROBERT ARTEAGA CASTRO",          # K - REPRESENTANTE EMPRESA
    "4800846G",                       # L - DNI REPRESENTANTE
    "ROBERT ARTEAGA CASTRO",          # M - TUTOR EMPRESA
    "4800846G",                       # N - DNI TUTOR
    "12/05/1982",                     # O - FECHA NACIMIENTO (Tutor)
    "info@kkubik.com"                 # P - CORREO EMPRESA
]

if __name__ == "__main__":
    # Nombre del archivo Excel de destino
    excel_file = "PATRON-CONVENNIOS.xlsx"
    
    # Ejecución de la actualización
    resultado = update_excel_with_extracted_data(excel_file, data_to_insert)
    print(resultado)